"""
excel_reporter.py — Generador de reportes en Excel.
"""

import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR = Path(__file__).parent.parent / "data"
REPORT_JSON_PATH = DATA_DIR / "latest_comparison_report.json"

def export_to_excel(json_path: Path = REPORT_JSON_PATH, output_path: Path | None = None) -> Path:
    if output_path is None:
        output_path = DATA_DIR / "reporte_diferencias_conaprole.xlsx"

    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    ws = wb.active
    ws.title = "Diferencias"

    # Estilos
    header_fill = PatternFill(start_color="2F225B", end_color="2F225B", fill_type="solid") # Morado Conaprole
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    red_fill = PatternFill(start_color="FFD1D1", end_color="FFD1D1", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    
    border_thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC")
    )

    headers = [
        "Nivel Alerta",
        "Supermercado",
        "Categoría",
        "Nombre Oficial Conaprole",
        "Nombre Publicado Supermercado",
        "Similitud Nombre (%)",
        "Estado Imagen",
        "Distancia pHash",
        "URL Imagen Conaprole",
        "URL Imagen Supermercado",
        "URL Producto Supermercado"
    ]

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for item in data.get("discrepancies", []):
        alert = item.get("alert_level", "YELLOW")
        sp_name = item.get("supermarket", "")
        off_prod = item.get("conaprole_product", {})
        sp_prod = item.get("supermarket_product", {})
        name_cmp = item.get("name_comparison", {})
        img_cmp = item.get("image_comparison") or {}

        alert_str = "🔴 APÓCRIFA" if alert == "RED" else "🟡 DIFERENTE"
        img_status = img_cmp.get("status", "SIN_DATOS") if img_cmp else "SIN_DATOS"
        phash_dist = img_cmp.get("phash_distance", "-") if img_cmp else "-"

        row = [
            alert_str,
            sp_name,
            off_prod.get("category", ""),
            off_prod.get("name", ""),
            sp_prod.get("name", ""),
            name_cmp.get("similarity_score", 0),
            img_status,
            phash_dist,
            off_prod.get("image_url", ""),
            sp_prod.get("image_url", ""),
            sp_prod.get("url", "")
        ]

        ws.append(row)
        current_row = ws.max_row
        
        # Formato de celda de alerta
        alert_cell = ws.cell(row=current_row, column=1)
        alert_cell.fill = red_fill if alert == "RED" else yellow_fill
        alert_cell.alignment = Alignment(horizontal="center")

        for col_idx in range(1, len(row) + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = border_thin

    # Autoajustar ancho de columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)

    wb.save(output_path)
    return output_path

if __name__ == "__main__":
    export_to_excel()
